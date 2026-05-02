"""
Export helpers — generate downloadable files from issues + raw data.

Three formats are supported:
- issues.csv      : flat list of all cell issues
- wrong_rows.csv  : only the rows that contain at least one invalid cell
- validation.xlsx : two-sheet workbook (annotated data + issues list)
"""
import io
import csv
import pandas as pd
from core.stats import get_invalid_row_set


def issues_to_csv(issues: dict) -> bytes:
    """
    Serialise all cell issues to a UTF-8 CSV with columns:
    row, column, code, message.
    """
    output = io.StringIO()
    writer = csv.DictWriter(
        output,
        fieldnames=["row", "column", "code", "message"],
        extrasaction="ignore",
        lineterminator="\n",
    )
    writer.writeheader()
    for ci in issues.get("cell_issues", []):
        writer.writerow(ci)
    return output.getvalue().encode("utf-8")


def wrong_rows_to_csv(df: pd.DataFrame, issues: dict) -> bytes:
    """
    Return a UTF-8 CSV containing only the rows that have at least one
    cell issue.  Preserves the original column order.
    """
    invalid_rows = get_invalid_row_set(issues)
    wrong_df     = df.iloc[sorted(invalid_rows)] if invalid_rows else df.iloc[[]]
    return wrong_df.to_csv(index=False).encode("utf-8")


def issues_to_xlsx(df: pd.DataFrame, issues: dict) -> bytes:
    """
    Build a two-sheet Excel workbook:

    Sheet 1 "Data"
        All rows from *df*.  Invalid cells are highlighted in red and carry an
        openpyxl Comment with the error code and message.  The header row is
        bold + blue-tinted, and the pane is frozen at row 2.

    Sheet 2 "Issues"
        Flat table of every cell issue (row / column / code / message) with a
        bold header row.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment

    # Build an error index keyed by (df_row_index, column_name).
    # Only the first error per cell is stored — this is sufficient for the
    # tooltip comment; the full list is available in Sheet 2.
    error_index: dict[tuple, tuple] = {}
    for ci in issues.get("cell_issues", []):
        key = (ci["row"], ci["column"])
        error_index.setdefault(key, (ci["code"], ci["message"]))

    wb = Workbook()

    # ── Sheet 1: annotated data ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    RED_FILL    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    BOLD        = Font(bold=True)

    columns = list(df.columns)

    # Write header row
    for col_idx, col_name in enumerate(columns, 1):
        cell      = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = BOLD
        cell.fill = HEADER_FILL

    # Write data rows.
    # Use df.iterrows() instead of itertuples() because itertuples mangles
    # column names that contain spaces or are not valid Python identifiers,
    # causing getattr() to silently return "" for those columns.
    for row_idx, (_, pandas_row) in enumerate(df.iterrows()):
        for col_idx, col_name in enumerate(columns, 1):
            val  = pandas_row[col_name]
            cell = ws.cell(
                row=row_idx + 2,
                column=col_idx,
                value=str(val) if val is not None else "",
            )
            err = error_index.get((row_idx, col_name))
            if err:
                cell.fill = RED_FILL
                try:
                    cell.comment = Comment(f"{err[0]}: {err[1]}", "DataValidator")
                except Exception:
                    pass  # Comments can fail with older openpyxl — non-fatal

    # Auto-fit column widths, capped at 40 characters
    for col_idx, col_name in enumerate(columns, 1):
        max_len = max(
            len(str(col_name)),
            *(len(str(ws.cell(row=r, column=col_idx).value or ""))
              for r in range(2, min(ws.max_row + 1, 102))),
        )
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"   # keep the header row visible when scrolling

    # ── Sheet 2: issues list ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Issues")

    # Write header row — bold each cell individually because setting
    # RowDimension.font is silently ignored by openpyxl.
    header_values = ["row", "column", "code", "message"]
    for col_idx, header in enumerate(header_values, 1):
        cell      = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD

    for ci in issues.get("cell_issues", []):
        ws2.append([
            ci.get("row",     ""),
            ci.get("column",  ""),
            ci.get("code",    ""),
            ci.get("message", ""),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
